"""엑셀 후공정 — 카드매입 세부내역을 카드사별로 묶고 합계행(수식)을 넣어 .xlsx로 저장.

- process_file(src, out): 단일 .xls를 카드사별 정렬 + 합계행 → .xlsx ('원본' 시트 보존)
- process_monthly_files(srcs, filter_from, filter_to, out): 여러 .xls를 합쳐
  '거래일자'(승인일/카드 쓴 날)가 [filter_from, filter_to]인 행만 남기고 카드사별 정렬.

    python postprocess.py <원본.xls> [출력.xlsx]
"""
from __future__ import annotations

import sys
from pathlib import Path

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TXN_DATE_COL = 1                          # 거래일자(승인일) 열 (0-based)
CARD_COL = 4                              # 카드사 열
AMT_COL = 8                               # 매입금액 열 (건수합계 표시용)
SUM_COLS = [8, 9, 10, 12, 13, 14, 15]     # 합계 낼 금액 열들 (수수료포인트률[11] 제외)

HEAD_FILL = PatternFill("solid", fgColor="E2E8F0")
SUB_FILL = PatternFill("solid", fgColor="FFF3C4")
BOLD = Font(bold=True)
RIGHT = Alignment(horizontal="right")


def read_xls_rows(path):
    sh = xlrd.open_workbook(str(path)).sheet_by_index(0)
    return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]


def find_header(rows):
    for i, row in enumerate(rows):
        if str(row[0]).strip() == "No.":
            return i
    return 2


def _norm(d):
    """YYYYMMDD/기타 → 'YYYY-MM-DD' (거래일자 문자열 비교용)."""
    s = "".join(ch for ch in str(d) if ch.isdigit())
    return f"{s[0:4]}-{s[4:6]}-{s[6:8]}" if len(s) >= 8 else str(d)


def write_report(out, title_rows, header, data, original_rows):
    """카드사별정렬 시트 + 원본 시트로 저장."""
    out = Path(out)
    ncol = len(header)
    wb = Workbook()
    ws = wb.active
    ws.title = "카드사별정렬"

    out_r = 1
    for tr in title_rows:
        for c in range(ncol):
            ws.cell(out_r, c + 1, tr[c] if c < len(tr) else None)
        out_r += 1
    for c in range(ncol):
        cell = ws.cell(out_r, c + 1, header[c])
        cell.font = BOLD
        cell.fill = HEAD_FILL
    out_r += 1
    first_data_row = out_r

    data = sorted(data, key=lambda r: str(r[CARD_COL]))
    i = 0
    while i < len(data):
        card = str(data[i][CARD_COL])
        start = out_r
        while i < len(data) and str(data[i][CARD_COL]) == card:
            for c in range(ncol):
                v = data[i][c] if c < len(data[i]) else None
                ws.cell(out_r, c + 1, v if v != "" else None)
            i += 1
            out_r += 1
        end = out_r - 1
        for c in range(ncol):
            ws.cell(out_r, c + 1).fill = SUB_FILL
            ws.cell(out_r, c + 1).font = BOLD
        ws.cell(out_r, CARD_COL + 1, f"{card} 합계")
        for col0 in SUM_COLS:
            L = get_column_letter(col0 + 1)
            ws.cell(out_r, col0 + 1).value = f"=SUM({L}{start}:{L}{end})"
        out_r += 1

    for col0 in SUM_COLS + [11]:
        for rr in range(first_data_row, out_r):
            ws.cell(rr, col0 + 1).number_format = "#,##0"
            ws.cell(rr, col0 + 1).alignment = RIGHT

    widths = {1: 5, 2: 12, 3: 12, 4: 10, 5: 11, 7: 14, 8: 9}
    for c in range(1, ncol + 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(c, 11)

    ws2 = wb.create_sheet("원본")
    for r in original_rows:
        ws2.append([(x if x != "" else None) for x in r])

    wb.save(out)
    return out


def read_detail(src):
    """원본 .xls의 헤더 + 세부내역 데이터행만 반환(정렬 안 함). 합본용."""
    rows = read_xls_rows(src)
    hi = find_header(rows)
    header = rows[hi]
    data = [r for r in rows[hi + 1:] if str(r[0]).strip() != ""]
    return header, data


def process_file(src, out=None):
    src = Path(src)
    if out is None:
        out = src.with_name(src.stem + "_정리.xlsx")
    rows = read_xls_rows(src)
    hi = find_header(rows)
    header = rows[hi]
    title_rows = rows[:hi]
    data = [r for r in rows[hi + 1:] if str(r[0]).strip() != ""]
    write_report(out, title_rows, header, data, rows)
    return header, data


def process_monthly_files(srcs, filter_from, filter_to, out, label=""):
    """여러 .xls를 합쳐 거래일자가 [filter_from, filter_to]인 행만 남기고 카드사별 정렬."""
    f_from, f_to = _norm(filter_from), _norm(filter_to)
    header = None
    merged = []
    for src in srcs:
        rows = read_xls_rows(src)
        hi = find_header(rows)
        if header is None:
            header = rows[hi]
        for r in rows[hi + 1:]:
            if str(r[0]).strip() == "":
                continue
            txn = _norm(r[TXN_DATE_COL])
            if f_from <= txn <= f_to:          # 거래일자 필터
                merged.append(r)
    if header is None:
        raise ValueError("헤더를 찾지 못했습니다(빈 파일?).")

    total_amt = sum((r[AMT_COL] for r in merged if isinstance(r[AMT_COL], (int, float))), 0)
    title_rows = [
        [f"기간별 매입내역 - 거래일 {label} 기준 정리"],
        ["건수", str(len(merged)), "합계", f"{int(total_amt):,}"],
    ]
    original_rows = [header] + merged
    write_report(out, title_rows, header, merged, original_rows)
    return header, merged


RESULT_FILL = {
    "성공": PatternFill("solid", fgColor="DCFCE7"),
    "조회없음": PatternFill("solid", fgColor="FEF9C3"),
    "로그인실패": PatternFill("solid", fgColor="FEE2E2"),
    "실패": PatternFill("solid", fgColor="FECACA"),
}


def build_results_report(results, out):
    """실행 결과 리포트. results: [(업체명, 결과라벨, 상세), ...].
    한 시트에 업체명/결과/상세, 결과는 색으로 구분."""
    out = Path(out)
    wb = Workbook()
    ws = wb.active
    ws.title = "실행결과"
    ws.append(["업체명", "결과", "상세"])
    for c in range(1, 4):
        ws.cell(1, c).font = BOLD
        ws.cell(1, c).fill = HEAD_FILL
    r = 2
    for (name, label, detail) in results:
        ws.cell(r, 1, name)
        cell = ws.cell(r, 2, label)
        cell.fill = RESULT_FILL.get(label, RESULT_FILL["실패"])
        cell.font = BOLD
        cell.alignment = Alignment(horizontal="center")
        ws.cell(r, 3, detail)
        r += 1
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 50
    ws.freeze_panes = "A2"
    wb.save(out)
    return out


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("usage: postprocess.py <원본.xls> [출력.xlsx]")
    else:
        o = process_file(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
        print("저장:", o)
