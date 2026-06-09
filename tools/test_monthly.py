# -*- coding: utf-8 -*-
"""월조회 후공정(합치기+거래일자 필터) 검증 — 데스크탑의 5월 파일 하나로 테스트."""
import io
import sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import postprocess

src = r"C:\Users\777la\OneDrive\바탕 화면\한국보건안전대구교육원_260501~260531.xls"
if not Path(src).exists():
    cands = []
    for d in (Path.home() / "OneDrive" / "바탕 화면", Path.home() / "Desktop", Path.home() / "Downloads"):
        if d.exists():
            cands += list(d.glob("*.xls"))
    src = str(max(cands, key=lambda p: p.stat().st_mtime)) if cands else src
print("src:", src)

# 원본 거래일자 분포 확인
rows = postprocess.read_xls_rows(src)
hi = postprocess.find_header(rows)
data = [r for r in rows[hi + 1:] if str(r[0]).strip() != ""]
from collections import Counter
months = Counter(postprocess._norm(r[postprocess.TXN_DATE_COL])[:7] for r in data)
print("원본 거래일자 월별 분포:", dict(months), " 총", len(data))

out = Path("tools/test_monthly_정리.xlsx")
postprocess.process_monthly_files([src], filter_from="20260501", filter_to="20260531",
                                  out=out, label="2026-05")
print("저장:", out)

import openpyxl
wb = openpyxl.load_workbook(out)
ws = wb["카드사별정렬"]
print("정리시트 행수:", ws.max_row)
# 거래일자(2열) 월 분포 확인 (헤더/합계행 제외하고 데이터행만)
m2 = Counter()
for r in range(4, ws.max_row + 1):
    v = ws.cell(r, 2).value
    if v and "-" in str(v):
        m2[str(v)[:7]] += 1
print("정리 후 거래일자 월별 분포:", dict(m2))
print("원본시트 행수:", wb["원본"].max_row)
